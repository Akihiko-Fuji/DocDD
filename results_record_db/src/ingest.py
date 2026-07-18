#!/usr/bin/python3
# -*- coding: utf-8 -*-
###########################################################################
# Name    : ingest.py
# Author  : Akihiko Fujita
# Update  : 2026-05-20
############################################################################

"""results_record_db のファイル取り込みロジックを実装するモジュール。

主な責務:
- CSV/XLSX の入力データを検証・正規化して取り込み計画を作成
- 業務ルールに基づくエラー判定とリジェクト理由の付与
- WorkLog / WorkLogReject への登録処理と取り込み集計結果の生成

"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - 任意依存のため未導入を許容
    load_workbook = None

from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session

from db import WorkLog, WorkLogReject

FILENAME_WORKER_RE = re.compile(
    r"^(INTASM)_([^_]+)_\d{6}(?:\d{2})?\.(csv|xlsx|xlsm)$"
    r"|^(EXTASM)(?:_[^_]+)?_([^_]+)_\d{6}(?:\d{2})?\.(csv|xlsx|xlsm)$"
)
HOLIDAYS = {
    date(2026, 1, 12)
}  # セミナー向け簡易実装: 土日+固定祝日(2026-01-12)のみを非稼働日として扱う。
REJECT_MISSING_REQUIRED = "MISSING_REQUIRED"
REJECT_DATE_PARSE_ERROR = "DATE_PARSE_ERROR"
REJECT_END_BEFORE_START = "END_BEFORE_START"
REJECT_WORK_EXCEEDS_ELAPSED = "WORK_EXCEEDS_ELAPSED"
REJECT_INVALID_RESULT_CD = "INVALID_RESULT_CD"
REJECT_ERROR_CODE_PRESENT = "ERROR_CODE_PRESENT"
REJECT_INVALID_WORKER_NAME = "INVALID_WORKER_NAME"
REJECT_DUPLICATE_KEY = "DUPLICATE_KEY"
REJECT_MASTER_NOT_FOUND = "MASTER_NOT_FOUND"
REJECT_DB_CONSTRAINT_ERROR = "DB_CONSTRAINT_ERROR"

PROCESS_BY_PREFIX = {
    "INTASM": ("内装組立", "internal_assembly_tool"),
    "EXTASM": ("外装組立", "external_assembly_tool"),
}
# セミナー専用互換: 旧教材サンプル名を許容するための暫定ホワイトリスト。
# 本番転用時は削除し、SHIPCHK_* 命名規則のみを受け付けること。
LEGACY_SHIPPING_SAMPLE_NAMES = {
    "shipping_inspection_log.csv",
    "shipping_inspection_log_invalid.csv",
}


@dataclass
class RejectRecord:
    source_system: Optional[str]
    source_file_name: str
    source_row_no: int
    reject_reason_cd: str
    reject_reason_detail: str
    raw_payload_json: str
    ingest_batch_id: str


@dataclass
class IngestResult:
    source_file_name: str
    ingest_batch_id: str
    inserted: int
    rejected: int


@dataclass
class NormalizedWorkLogCandidate:
    order_no: str
    product_name: str
    process_name: str
    worker_name: str
    start_ts: datetime
    end_ts: datetime
    elapsed_sec: int
    work_sec: int
    result_cd: str
    source_system: str
    source_file_name: str
    source_row_no: int
    ingest_batch_id: str
    raw_payload_json: str


@dataclass
class IngestPlan:
    source_file_name: str
    ingest_batch_id: str
    candidates: List[NormalizedWorkLogCandidate]
    rejects: List[RejectRecord]


@dataclass(frozen=True)
class SourceFileInfo:
    """ファイル名から判定したログ種別と作業者名。"""

    log_type: str
    file_worker: Optional[str]


@dataclass(frozen=True)
class ParsedWorkLog:
    """入力元固有の列を共通項目へ読み替えた1行分のデータ。"""

    order_no: str
    product_name: str
    process_name: str
    worker_name: str
    start_ts: datetime
    end_ts: datetime
    result_cd: str
    source_system: str


class RowValidationError(ValueError):
    """入力行を reject にする理由を呼び出し元へ伝える。"""

    def __init__(self, code: str, detail: str, source_system: str) -> None:
        super().__init__(detail)
        self.code = code
        self.detail = detail
        self.source_system = source_system


def load_order_product_master(path: Optional[str]) -> Dict[str, str]:
    """受注-製品マスタCSVを読み込み、必須列を検証して返す。"""
    if not path:
        return {}
    master_path = Path(path)
    if not master_path.exists():
        raise FileNotFoundError(f"order_product_master not found: {master_path}")

    with master_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required_columns = {"order_no", "product_name"}
        if not reader.fieldnames or not required_columns.issubset(
            set(reader.fieldnames)
        ):
            raise ValueError(
                "order_product_master.csv must contain order_no and product_name columns"
            )

        result: Dict[str, str] = {}
        for row in reader:
            order_no_raw = row.get("order_no")
            product_name_raw = row.get("product_name")
            order_no = "" if order_no_raw is None else str(order_no_raw).strip()
            product_name = (
                "" if product_name_raw is None else str(product_name_raw).strip()
            )
            if order_no and product_name:
                result[normalize_order_no(order_no)] = product_name
    return result


def normalize_worker_name(name: Optional[str]) -> str:
    """作業者名の空白を正規化して比較可能な文字列にする。"""
    if name is None:
        return ""
    return re.sub(r"[\s\u3000]+", "", str(name)).strip()


def normalize_order_no(order_no: Optional[str]) -> str:
    """受注番号の前後空白だけを除去する。

    ハイフンや英字の大文字・小文字は業務キーの一部として保持する。
    記号除去で異なる受注番号を同一視しないための仕様である。
    """
    if order_no is None:
        return ""
    return str(order_no).strip()


def calc_work_sec(start_ts: datetime, end_ts: datetime) -> int:
    """妥当な時刻範囲に対する所定勤務時間内の実作業秒数を計算する。"""
    if end_ts < start_ts:
        raise ValueError("end_ts must be greater than or equal to start_ts")

    total = 0
    cur_day = start_ts.date()
    end_day = end_ts.date()
    one_day = timedelta(days=1)

    while cur_day <= end_day:
        if cur_day.weekday() < 5 and cur_day not in HOLIDAYS:
            slots = [
                (
                    datetime.combine(cur_day, time(8, 0)),
                    datetime.combine(cur_day, time(12, 0)),
                ),
                (
                    datetime.combine(cur_day, time(13, 0)),
                    datetime.combine(cur_day, time(17, 0)),
                ),
            ]
            for slot_start, slot_end in slots:
                overlap_start = max(start_ts, slot_start)
                overlap_end = min(end_ts, slot_end)
                if overlap_end > overlap_start:
                    total += int((overlap_end - overlap_start).total_seconds())
        cur_day += one_day
    return total


def _parse_datetime(value: str, fmts: Sequence[str]) -> Optional[datetime]:
    """複数フォーマット候補で日時文字列を解析する。"""
    txt = str(value).strip()
    if not txt:
        return None
    for fmt in fmts:
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return None


def _missing_required(raw: Dict[str, str], columns: Sequence[str]) -> Optional[str]:
    """必須列のうち空値の最初の列名を返す。"""
    for col in columns:
        if not str(raw.get(col, "")).strip():
            return col
    return None


def _iter_input_rows(file_path: Path) -> Iterable[Tuple[int, Dict[str, str]]]:
    """入力ファイルを行番号付き辞書行に変換して列挙する。"""
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            header = [
                str(c).strip() if c is not None else ""
                for c in (reader.fieldnames or [])
            ]
            if (
                not header
                or any(not col for col in header)
                or len(set(header)) != len(header)
            ):
                raise ValueError(
                    "CSV header must be non-empty and contain unique column names"
                )
            # DictReader の辞書キーを検証後ヘッダーに統一し、strip 後ヘッダーとの乖離を防ぐ。
            reader.fieldnames = header
            for idx, row in enumerate(reader, start=1):
                yield idx, {k: (v if v is not None else "") for k, v in row.items()}
        return

    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ImportError("openpyxl is required to read .xlsx/.xlsm files")
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows)
            except StopIteration:
                return
            header = [str(c).strip() if c is not None else "" for c in first_row]
            if (
                not header
                or any(not col for col in header)
                or len(set(header)) != len(header)
            ):
                raise ValueError(
                    "Spreadsheet header must be non-empty and contain unique column names"
                )
            width = len(header)
            for idx, values in enumerate(rows, start=1):
                padded = list(values) + [None] * max(0, width - len(values))
                row = {
                    header[i]: ("" if padded[i] is None else str(padded[i]))
                    for i in range(width)
                }
                yield idx, row
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported file extension: {file_path.suffix}")


def _batch_id(seq: int = 1, override: Optional[str] = None) -> str:
    """README互換の ingest_batch_id を生成する。"""
    normalized_override = None if override is None else str(override).strip()
    if normalized_override:
        if len(normalized_override) > 30:
            raise ValueError("ingest_batch_id must be 30 characters or fewer")
        return normalized_override
    normalized_seq = max(seq, 1)
    # 仕様合意: ingest_batch_id は DB の String(30) に収める。
    # 固定長20文字(ING_ + YYYYMMDD_HHMMSS + _)のため、seq は9桁以下に制限する。
    if normalized_seq >= 1_000_000_000:
        raise ValueError(
            "ingest sequence is too large for ingest_batch_id length limit (max 9 digits)"
        )
    return f"ING_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{normalized_seq:03d}"


def _next_business_day(d: date) -> date:
    """土日・固定祝日を除いた翌営業日を返す。"""
    cur = d + timedelta(days=1)
    while cur.weekday() >= 5 or cur in HOLIDAYS:
        cur += timedelta(days=1)
    return cur


def _detect_source_file(file_name: str) -> SourceFileInfo:
    """ファイル名からログ種別とファイル名由来の作業者名を判定する。"""
    match = FILENAME_WORKER_RE.match(file_name)
    file_prefix = None
    file_worker = None
    if match:
        if match.group(1):
            file_prefix = match.group(1)
            file_worker = normalize_worker_name(match.group(2))
        else:
            file_prefix = match.group(4)
            file_worker = normalize_worker_name(match.group(5))

    normalized_name = file_name.lower()
    is_shipping = (
        file_name.startswith("SHIPCHK_")
        or normalized_name in LEGACY_SHIPPING_SAMPLE_NAMES
    )
    if is_shipping:
        return SourceFileInfo("shipping", None)
    if file_prefix == "INTASM":
        return SourceFileInfo("internal", file_worker)
    if file_prefix == "EXTASM":
        return SourceFileInfo("external", file_worker)
    raise ValueError(f"Unknown file naming: {file_name}")


def _resolve_worker_name(
    file_worker: Optional[str], row_worker: Optional[str], source_system: str
) -> str:
    """ファイル名と行データから作業者名を解決し、不一致を検出する。"""
    normalized_row_worker = normalize_worker_name(row_worker)
    if file_worker and normalized_row_worker and file_worker != normalized_row_worker:
        raise RowValidationError(
            REJECT_INVALID_WORKER_NAME,
            "worker_name mismatch between file name and row",
            source_system,
        )
    worker_name = normalize_worker_name(file_worker or normalized_row_worker)
    if not worker_name:
        raise RowValidationError(
            REJECT_INVALID_WORKER_NAME,
            "worker_name could not be resolved",
            source_system,
        )
    return worker_name


def _parse_internal_row(
    raw: Dict[str, str],
    file_worker: Optional[str],
    order_product_map: Dict[str, str],
) -> ParsedWorkLog:
    """内装組立ログ1行を共通項目へ変換する。"""
    process_name, source_system = PROCESS_BY_PREFIX["INTASM"]
    order_no = normalize_order_no(raw.get("order_no"))
    if not order_no:
        raise RowValidationError(
            REJECT_MISSING_REQUIRED, "order_no is empty", source_system
        )

    missing = _missing_required(
        raw, ["start_date", "start_time", "end_date", "end_time"]
    )
    if missing:
        raise RowValidationError(
            REJECT_MISSING_REQUIRED, f"{missing} is empty", source_system
        )
    if str(raw.get("start_marker", "")).strip() != "START":
        raise RowValidationError(
            REJECT_MISSING_REQUIRED, "start_marker must be START", source_system
        )
    if str(raw.get("end_marker", "")).strip() != "END":
        raise RowValidationError(
            REJECT_MISSING_REQUIRED, "end_marker must be END", source_system
        )

    start_ts = _parse_datetime(
        f"{raw['start_date']} {raw['start_time']}",
        ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"],
    )
    end_ts = _parse_datetime(
        f"{raw['end_date']} {raw['end_time']}",
        ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"],
    )
    if not start_ts or not end_ts:
        raise RowValidationError(
            REJECT_DATE_PARSE_ERROR,
            "failed to parse start/end timestamp",
            source_system,
        )

    product_name = str(order_product_map.get(order_no, "")).strip()
    if not product_name:
        raise RowValidationError(
            REJECT_MASTER_NOT_FOUND,
            "product_name not found by order_no",
            source_system,
        )

    return ParsedWorkLog(
        order_no=order_no,
        product_name=product_name,
        process_name=process_name,
        worker_name=_resolve_worker_name(
            file_worker, raw.get("worker_name"), source_system
        ),
        start_ts=start_ts,
        end_ts=end_ts,
        result_cd="OK",
        source_system=source_system,
    )


def _parse_external_row(
    raw: Dict[str, str], file_worker: Optional[str]
) -> ParsedWorkLog:
    """外装組立ログ1行を共通項目へ変換する。"""
    process_name, source_system = PROCESS_BY_PREFIX["EXTASM"]
    order_no = normalize_order_no(raw.get("order_no"))
    product_name = str(raw.get("product_name", "")).strip()
    if not order_no or not product_name:
        raise RowValidationError(
            REJECT_MISSING_REQUIRED,
            "order_no or product_name is empty",
            source_system,
        )
    if not str(raw.get("qr_read_ts", "")).strip() or not str(
        raw.get("all_clear_ts", "")
    ).strip():
        raise RowValidationError(
            REJECT_MISSING_REQUIRED,
            "qr_read_ts or all_clear_ts is empty",
            source_system,
        )

    start_ts = _parse_datetime(raw["qr_read_ts"], ["%Y-%m-%d %H:%M:%S"])
    end_ts = _parse_datetime(raw["all_clear_ts"], ["%Y-%m-%d %H:%M:%S"])
    if not start_ts or not end_ts:
        raise RowValidationError(
            REJECT_DATE_PARSE_ERROR,
            "failed to parse qr_read_ts/all_clear_ts",
            source_system,
        )
    if str(raw.get("error_code", "")).strip():
        raise RowValidationError(
            REJECT_ERROR_CODE_PRESENT, "error_code must be empty", source_system
        )

    return ParsedWorkLog(
        order_no=order_no,
        product_name=product_name,
        process_name=process_name,
        worker_name=_resolve_worker_name(
            file_worker, raw.get("worker_name"), source_system
        ),
        start_ts=start_ts,
        end_ts=end_ts,
        result_cd="OK",
        source_system=source_system,
    )


def _parse_shipping_row(raw: Dict[str, str]) -> ParsedWorkLog:
    """出荷検査ログ1行を共通項目へ変換する。"""
    process_name = "出荷検査"
    source_system = "shipping_inspection_tool"
    order_no = normalize_order_no(raw.get("order_no"))
    product_name = str(raw.get("product_name", "")).strip()
    worker_name = normalize_worker_name(raw.get("inspector_name"))
    if not order_no or not product_name or not worker_name:
        raise RowValidationError(
            REJECT_MISSING_REQUIRED,
            "order_no/product_name/inspector_name is empty",
            source_system,
        )

    missing = _missing_required(
        raw, ["inspection_date", "start_time", "end_time", "ng_total"]
    )
    if missing:
        raise RowValidationError(
            REJECT_MISSING_REQUIRED, f"{missing} is empty", source_system
        )

    start_ts = _parse_datetime(
        f"{raw['inspection_date']} {raw['start_time']}", ["%Y-%m-%d %H:%M:%S"]
    )
    end_ts = _parse_datetime(
        f"{raw['inspection_date']} {raw['end_time']}", ["%Y-%m-%d %H:%M:%S"]
    )
    if not start_ts or not end_ts:
        raise RowValidationError(
            REJECT_DATE_PARSE_ERROR,
            "failed to parse inspection timestamp",
            source_system,
        )
    if end_ts < start_ts:
        end_ts = datetime.combine(_next_business_day(start_ts.date()), end_ts.time())

    try:
        ng_total = int(str(raw["ng_total"]).strip())
    except ValueError as exc:
        raise RowValidationError(
            REJECT_INVALID_RESULT_CD, "ng_total is not integer", source_system
        ) from exc

    return ParsedWorkLog(
        order_no=order_no,
        product_name=product_name,
        process_name=process_name,
        worker_name=worker_name,
        start_ts=start_ts,
        end_ts=end_ts,
        result_cd="OK" if ng_total == 0 else "NG",
        source_system=source_system,
    )


def _parse_source_row(
    source: SourceFileInfo,
    raw: Dict[str, str],
    order_product_map: Dict[str, str],
) -> ParsedWorkLog:
    """判定済みログ種別に対応する行変換処理を呼び出す。"""
    if source.log_type == "internal":
        return _parse_internal_row(raw, source.file_worker, order_product_map)
    if source.log_type == "external":
        return _parse_external_row(raw, source.file_worker)
    return _parse_shipping_row(raw)


def _load_existing_keys(
    session: Session, rows: Sequence[Tuple[int, Dict[str, str]]]
) -> set[tuple[str, str]]:
    """入力候補の受注番号に対応する既存の業務キーだけを取得する。"""
    candidate_order_nos = set()
    for _, raw in rows:
        order_no = normalize_order_no(raw.get("order_no"))
        if order_no:
            candidate_order_nos.add(order_no)
    existing_keys: set[tuple[str, str]] = set()
    order_no_list = sorted(candidate_order_nos)
    for start in range(0, len(order_no_list), 900):
        chunk = order_no_list[start : start + 900]
        existing_keys.update(
            (row.order_no, row.process_name)
            for row in session.query(WorkLog.order_no, WorkLog.process_name)
            .filter(WorkLog.order_no.in_(chunk))
            .all()
        )
    return existing_keys


def _append_reject(
    rejects: List[RejectRecord],
    *,
    source_system: Optional[str],
    source_file_name: str,
    source_row_no: int,
    reject_reason_cd: str,
    reject_reason_detail: str,
    raw_payload: Dict[str, str],
    ingest_batch_id: str,
) -> None:
    """reject レコードを1件追加する。"""
    rejects.append(
        RejectRecord(
            source_system=source_system,
            source_file_name=source_file_name,
            source_row_no=source_row_no,
            reject_reason_cd=reject_reason_cd,
            reject_reason_detail=reject_reason_detail,
            raw_payload_json=json.dumps(raw_payload, ensure_ascii=False),
            ingest_batch_id=ingest_batch_id,
        )
    )


def prepare_ingest_file(
    session: Session,
    file_path: str,
    order_product_map: Optional[Dict[str, str]] = None,
    ingest_batch_id: Optional[str] = None,
    ingest_seq: int = 1,
) -> IngestPlan:
    """単一ファイルを検証し、DB登録候補とreject候補へ分ける。"""
    master = order_product_map or {}
    file_path_obj = Path(file_path)
    batch_id = _batch_id(ingest_seq, override=ingest_batch_id)
    source = _detect_source_file(file_path_obj.name)

    rows = list(_iter_input_rows(file_path_obj))
    existing_keys = _load_existing_keys(session, rows)
    seen_keys: set[tuple[str, str]] = set()
    candidates: List[NormalizedWorkLogCandidate] = []
    rejects: List[RejectRecord] = []

    for row_no, raw in rows:
        try:
            parsed = _parse_source_row(source, raw, master)

            # 出荷検査の日跨ぎは個別変換時に補正済み。ここでは全ログ共通の
            # 時刻・作業時間・重複ルールだけを評価する。
            if parsed.end_ts < parsed.start_ts:
                raise RowValidationError(
                    REJECT_END_BEFORE_START,
                    "end_ts before start_ts",
                    parsed.source_system,
                )
            elapsed_sec = int((parsed.end_ts - parsed.start_ts).total_seconds())
            work_sec = calc_work_sec(parsed.start_ts, parsed.end_ts)
            if work_sec > elapsed_sec:
                raise RowValidationError(
                    REJECT_WORK_EXCEEDS_ELAPSED,
                    "work_sec > elapsed_sec",
                    parsed.source_system,
                )

            dedup_key = (parsed.order_no, parsed.process_name)
            if dedup_key in existing_keys or dedup_key in seen_keys:
                raise RowValidationError(
                    REJECT_DUPLICATE_KEY,
                    "UNIQUE(order_no, process_name) violation",
                    parsed.source_system,
                )
        except RowValidationError as exc:
            _append_reject(
                rejects,
                source_system=exc.source_system,
                source_file_name=file_path_obj.name,
                source_row_no=row_no,
                reject_reason_cd=exc.code,
                reject_reason_detail=exc.detail,
                raw_payload=raw,
                ingest_batch_id=batch_id,
            )
            continue

        seen_keys.add(dedup_key)
        candidates.append(
            NormalizedWorkLogCandidate(
                order_no=parsed.order_no,
                product_name=parsed.product_name,
                process_name=parsed.process_name,
                worker_name=parsed.worker_name,
                start_ts=parsed.start_ts,
                end_ts=parsed.end_ts,
                elapsed_sec=elapsed_sec,
                work_sec=work_sec,
                result_cd=parsed.result_cd,
                source_system=parsed.source_system,
                source_file_name=file_path_obj.name,
                source_row_no=row_no,
                ingest_batch_id=batch_id,
                raw_payload_json=json.dumps(raw, ensure_ascii=False),
            )
        )

    return IngestPlan(
        source_file_name=file_path_obj.name,
        ingest_batch_id=batch_id,
        candidates=candidates,
        rejects=rejects,
    )


def _work_log_from_candidate(candidate: NormalizedWorkLogCandidate) -> WorkLog:
    """正規化済み候補をORMモデルへ変換する。"""
    return WorkLog(
        order_no=candidate.order_no,
        product_name=candidate.product_name,
        process_name=candidate.process_name,
        worker_name=candidate.worker_name,
        start_ts=candidate.start_ts,
        end_ts=candidate.end_ts,
        elapsed_sec=candidate.elapsed_sec,
        work_sec=candidate.work_sec,
        result_cd=candidate.result_cd,
        source_system=candidate.source_system,
        source_file_name=candidate.source_file_name,
        source_row_no=candidate.source_row_no,
        ingest_batch_id=candidate.ingest_batch_id,
    )


def _reject_from_candidate(
    candidate: NormalizedWorkLogCandidate, code: str, detail: str
) -> RejectRecord:
    """DB登録に失敗した候補からrejectレコードを作る。"""
    return RejectRecord(
        source_system=candidate.source_system,
        source_file_name=candidate.source_file_name,
        source_row_no=candidate.source_row_no,
        reject_reason_cd=code,
        reject_reason_detail=detail,
        raw_payload_json=candidate.raw_payload_json,
        ingest_batch_id=candidate.ingest_batch_id,
    )


def _is_duplicate_integrity_error(exc: IntegrityError) -> bool:
    """PostgreSQL / SQLite の一意制約違反を同じ基準で判定する。"""
    original = getattr(exc, "orig", None)
    detail = str(original or exc).lower()
    sqlstate = getattr(original, "pgcode", None)
    return (
        sqlstate == "23505"
        or "duplicate key" in detail
        or "unique constraint failed: work_log.order_no, work_log.process_name"
        in detail
        or "uq_work_log_order_no_process_name" in detail
    )


def _persist_reject(session: Session, reject: RejectRecord) -> None:
    """rejectレコードをORMモデルへ変換してセッションへ追加する。"""
    session.add(
        WorkLogReject(
            source_system=reject.source_system,
            source_file_name=reject.source_file_name,
            source_row_no=reject.source_row_no,
            reject_reason_cd=reject.reject_reason_cd,
            reject_reason_detail=reject.reject_reason_detail,
            raw_payload_json=reject.raw_payload_json,
            ingest_batch_id=reject.ingest_batch_id,
        )
    )


def apply_ingest_plan(session: Session, plan: IngestPlan) -> IngestResult:
    """取り込み計画をDBへ反映し、結果件数を返す。"""
    inserted = 0
    rejects = list(plan.rejects)
    for candidate in plan.candidates:
        try:
            with session.begin_nested():
                session.add(_work_log_from_candidate(candidate))
                session.flush()
                inserted += 1
        except IntegrityError as exc:
            is_duplicate = _is_duplicate_integrity_error(exc)
            code = REJECT_DUPLICATE_KEY if is_duplicate else REJECT_DB_CONSTRAINT_ERROR
            detail = (
                "UNIQUE(order_no, process_name) violation"
                if is_duplicate
                else str(getattr(exc, "orig", None) or exc)
            )
            rejects.append(_reject_from_candidate(candidate, code, detail))
        except SQLAlchemyError as exc:
            rejects.append(
                _reject_from_candidate(
                    candidate,
                    REJECT_DB_CONSTRAINT_ERROR,
                    str(exc),
                )
            )

    for reject in rejects:
        _persist_reject(session, reject)
    session.flush()
    session.commit()
    return IngestResult(
        source_file_name=plan.source_file_name,
        ingest_batch_id=plan.ingest_batch_id,
        inserted=inserted,
        rejected=len(rejects),
    )


def ingest_file(
    session: Session,
    file_path: str,
    order_product_map: Optional[Dict[str, str]] = None,
    ingest_batch_id: Optional[str] = None,
    ingest_seq: int = 1,
) -> IngestResult:
    """単一ファイルを前処理からDB反映まで一括で実行する。"""
    plan = prepare_ingest_file(
        session,
        file_path,
        order_product_map=order_product_map,
        ingest_batch_id=ingest_batch_id,
        ingest_seq=ingest_seq,
    )
    return apply_ingest_plan(session, plan)


def ingest_files(
    session: Session,
    file_paths: Sequence[str],
    order_product_map: Optional[Dict[str, str]] = None,
) -> List[IngestResult]:
    """複数ファイルを順次取り込み、各ファイル結果を返す。"""
    results: List[IngestResult] = []
    for seq, file_path in enumerate(file_paths, start=1):
        results.append(
            ingest_file(
                session=session,
                file_path=file_path,
                order_product_map=order_product_map,
                ingest_seq=seq,
            )
        )
    return results
